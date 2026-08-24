#!/usr/bin/env python3
"""
convert.py — Novara Business Intelligence
Converts the sales closure Excel workbook into data.json consumed by index.html.

Usage:
    python convert.py CLOSURE_DASHBOARD_new.xlsx            # writes data.json next to this script
    python convert.py my_file.xlsx -o public/data.json      # custom output path

The same cleaning rules are implemented in index.html (browser upload) so that
uploading the Excel in the DATA tab produces an identical data.json.

Requires: openpyxl  (pip install openpyxl)
"""
import argparse, datetime as dt, json, math, os, re, sys, warnings

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

warnings.filterwarnings("ignore")

# --------------------------------------------------------------------------
# Header-name detection (preferred over fixed column letters)
# --------------------------------------------------------------------------
FIELD_HEADERS = {
    # field        : list of acceptable header names (case/space-insensitive)
    "sno":          ["s.no", "sno", "sr no", "serial"],
    "agent":        ["agent name", "agent", "sales person", "salesperson", "consultant"],
    "date":         ["date of closure", "closure date", "booking date", "date", "sales date", "transaction date"],
    "property":     ["property", "project", "project name"],
    "unit":         ["unit no", "unit", "unit number"],
    "developer":    ["developer", "developer name"],
    "client":       ["client name", "client", "customer", "customer name"],
    "rm":           ["rm", "relationship manager"],
    "source":       ["source", "lead source"],
    "mop":          ["mode of payment", "payment mode"],
    "tokenPaid":    ["token paid", "token"],
    "dpStatus":     ["dp status", "down payment status"],
    "docStatus":    ["doc status", "document status"],
    "propertyValue":["property value", "unit value", "sale value"],
    "commPct":      ["commission %", "commission percent"],
    "commPayout":   ["commission payout", "commission"],
    "passback":     ["passback", "pass back"],
    "vatInc":       ["vat (inclusive)", "vat inclusive"],
    "vatExc":       ["vat (exclusive)", "vat exclusive"],
    "gross":        ["gross commission", "gross"],
    "agentPct":     ["agent %"],
    "toAgent":      ["% to agent", "agent commission"],
    "tlPct":        ["tl %"],
    "toTL":         ["% to team leader"],
    "shPct":        ["sh %"],
    "toSH":         ["% to sales head"],
    "toTele":       ["% to telecaller"],
    "net":          ["net revenue for novara", "net revenue", "net"],
    "invoice":      ["invoice number", "invoice no", "invoice"],
    "invoiceDate":  ["invoice date"],
    "team":         ["team", "team name"],
    "year":         ["year"],
    "month":        ["month"],
}
REQUIRED = ["agent", "date", "developer", "net", "team"]

# Fallback letters (only used if a header is not found)
FALLBACK_LETTERS = {"agent": "B", "date": "C", "developer": "F", "net": "AB", "team": "AF"}

# Agent name aliases -> canonical (all UPPERCASE, trimmed). Edit freely.
AGENT_ALIASES = {
    "BAZEED KHAN": "BAZEED",
    "SHOAIB": "SHUAIB",
    "AWAIS (B2B)": "AWAIS",
    "DANSIH": "DANISH",
    "B2B (FURQAN)": "FURQAN",
    "B2B (TABISH)": "TABISH",
    "BILAL FAROOQ + LUV": "BILAL FAROOQ",
}

MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def norm_header(h):
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def col_letter(idx):
    s = ""
    idx += 1
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def letter_idx(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def clean_text(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def clean_number(v):
    """AED 25,000 -> 25000 ; (1,200) -> -1200 ; '' -> None"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s.replace(",", ""))
    if s in ("", "-", "."):
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def parse_date(v):
    """Returns datetime.date or None. Handles Excel serials, datetimes, common strings."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        if 20000 < v < 80000:  # Excel serial
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(v))).date()
        return None
    s = str(v).strip().splitlines()[0].strip() if str(v).strip() else ""
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y", "%b %d, %Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        if mo > 12 and d <= 12:  # MM/DD/YYYY
            d, mo = mo, d
        try:
            return dt.date(y, mo, d) if 1990 <= y <= 2100 else None
        except ValueError:
            return None
    return None


def month_index(v):
    s = clean_text(v)
    if not s:
        return None
    key = s.upper()[:3]
    return MONTHS.index(key) + 1 if key in MONTHS else None


def canonical_agent(raw):
    s = clean_text(raw)
    if not s:
        return None
    s = s.upper()
    return AGENT_ALIASES.get(s, s)


# --------------------------------------------------------------------------
def convert(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    # 1) find header row = first row containing "net revenue" or "agent"
    header_row = None
    for i, r in enumerate(rows[:15]):
        cells = [norm_header(c) for c in r]
        if any("net revenue" in c for c in cells) and any("agent" in c or "date" in c for c in cells):
            header_row = i
            break
    if header_row is None:
        sys.exit("Could not locate a header row containing 'Net Revenue for Novara'.")
    headers = [norm_header(c) for c in rows[header_row]]

    # 2) map fields by header name, fallback to letters
    colmap = {}
    for field, names in FIELD_HEADERS.items():
        for n in names:
            if n in headers:
                colmap[field] = headers.index(n)
                break
    for f, letter in FALLBACK_LETTERS.items():
        if f not in colmap and letter_idx(letter) < len(headers):
            colmap[f] = letter_idx(letter)
    missing = [f for f in REQUIRED if f not in colmap]
    if missing:
        sys.exit("Missing required columns: " + ", ".join(missing))

    def get(r, f):
        i = colmap.get(f)
        return r[i] if i is not None and i < len(r) else None

    q = dict(totalRows=0, blankRows=0, valid=0, missingDate=0, estimatedDate=0, missingAgent=0,
             missingDeveloper=0, missingTeam=0, invalidRevenue=0, zeroRevenue=0, duplicates=0, junkRows=0)
    seen = set()
    txns = []
    for r in rows[header_row + 1:]:
        q["totalRows"] += 1
        if all(c is None or str(c).strip() == "" for c in r):
            q["blankRows"] += 1
            continue
        agent = canonical_agent(get(r, "agent"))
        date = parse_date(get(r, "date"))
        net = clean_number(get(r, "net"))
        dev = clean_text(get(r, "developer"))
        team = clean_text(get(r, "team"))
        # junk row: no agent, no date, no revenue
        if agent is None and date is None and net is None:
            q["junkRows"] += 1
            continue
        est = False
        if date is None:
            y = clean_number(get(r, "year"))
            m = month_index(get(r, "month"))
            if y and m:
                date = dt.date(int(y), m, 1)
                est = True
                q["estimatedDate"] += 1
            else:
                q["missingDate"] += 1
        if agent is None:
            q["missingAgent"] += 1
        if dev is None:
            q["missingDeveloper"] += 1
        if team is None:
            q["missingTeam"] += 1
        if net is None:
            q["invalidRevenue"] += 1
            net = 0.0
        elif net == 0:
            q["zeroRevenue"] += 1
        key = (agent, date.isoformat() if date else None, dev, clean_text(get(r, "unit")), round(net, 2), clean_text(get(r, "client")))
        dup = key in seen
        if dup:
            q["duplicates"] += 1
        seen.add(key)
        inv_date = parse_date(get(r, "invoiceDate"))
        t = {
            "id": len(txns) + 1,
            "sno": clean_text(get(r, "sno")),
            "agent": agent or "UNASSIGNED",
            "agentRaw": clean_text(get(r, "agent")),
            "date": date.isoformat() if date else None,
            "dateEstimated": est,
            "property": clean_text(get(r, "property")),
            "unit": clean_text(get(r, "unit")),
            "developer": (dev or "UNASSIGNED").upper(),
            "client": clean_text(get(r, "client")),
            "rm": clean_text(get(r, "rm")),
            "source": (clean_text(get(r, "source")) or "").upper() or None,
            "mop": (clean_text(get(r, "mop")) or "").upper() or None,
            "tokenPaid": clean_number(get(r, "tokenPaid")),
            "dpStatus": clean_text(get(r, "dpStatus")),
            "docStatus": clean_text(get(r, "docStatus")),
            "propertyValue": clean_number(get(r, "propertyValue")),
            "commPct": clean_number(get(r, "commPct")),
            "commPayout": clean_number(get(r, "commPayout")),
            "passback": clean_number(get(r, "passback")),
            "vatInc": clean_number(get(r, "vatInc")),
            "vatExc": clean_number(get(r, "vatExc")),
            "gross": clean_number(get(r, "gross")),
            "toAgent": clean_number(get(r, "toAgent")),
            "toTL": clean_number(get(r, "toTL")),
            "toSH": clean_number(get(r, "toSH")),
            "toTele": clean_number(get(r, "toTele")),
            "net": math.floor(net * 100 + 0.5) / 100 if net >= 0 else -math.floor(-net * 100 + 0.5) / 100,
            "invoice": clean_text(get(r, "invoice")),
            "invoiceDate": inv_date.isoformat() if inv_date else None,
            "team": (team or "UNASSIGNED").upper(),
            "duplicate": dup,
        }
        txns.append(t)
        q["valid"] += 1

    # optional expense sheet(s): two-column pivot "Row Labels | Sum of ..."
    expenses = {}
    for sh in wb.worksheets[1:]:
        if re.fullmatch(r"20\d\d", sh.title.strip()):
            items = []
            for r in sh.iter_rows(values_only=True):
                if r and r[0] and len(r) > 1 and isinstance(r[1], (int, float)) and "total" not in str(r[0]).lower():
                    items.append({"category": clean_text(r[0]), "amount": round(float(r[1]), 2)})
            if items:
                expenses[sh.title.strip()] = items

    dates = [t["date"] for t in txns if t["date"]]
    return {
        "meta": {
            "app": "Novara Business Intelligence",
            "sourceFile": os.path.basename(path),
            "sheet": ws.title,
            "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
            "generatedBy": "convert.py",
            "columnMap": {f: {"index": i, "letter": col_letter(i), "header": rows[header_row][i]} for f, i in colmap.items()},
            "dateRange": {"from": min(dates) if dates else None, "to": max(dates) if dates else None},
            "quality": q,
            "agentAliases": AGENT_ALIASES,
        },
        "transactions": txns,
        "expenses": expenses,
    }


def main():
    ap = argparse.ArgumentParser(description="Convert Novara closure Excel to data.json")
    ap.add_argument("excel", help="Path to .xlsx workbook")
    ap.add_argument("-o", "--output", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json"))
    a = ap.parse_args()
    data = convert(a.excel)
    with open(a.output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    q = data["meta"]["quality"]
    tot = sum(t["net"] for t in data["transactions"])
    print(f"Wrote {a.output}")
    print(f"  transactions: {q['valid']}  | net revenue: AED {tot:,.2f}")
    print(f"  blank rows: {q['blankRows']}  junk: {q['junkRows']}  estimated dates: {q['estimatedDate']}  "
          f"missing developer: {q['missingDeveloper']}  missing team: {q['missingTeam']}  duplicates: {q['duplicates']}")
    print(f"  date range: {data['meta']['dateRange']['from']} -> {data['meta']['dateRange']['to']}")


if __name__ == "__main__":
    main()
